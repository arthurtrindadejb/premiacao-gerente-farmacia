import io
import json
import os
from datetime import date

from flask import Flask, abort, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_wtf import CSRFProtect
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from werkzeug.security import generate_password_hash

from auth import admin_required, login_required, verificar_login
from calculo import atingimento, calcular_mes
from db import db

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ["SECRET_KEY"]
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "1") == "1"
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
csrf = CSRFProtect(app)

MESES_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def mes_label(mes):
    ano, mo = mes.split("-")
    return f"{MESES_PT[int(mo) - 1]} / {ano}"


def carregar_gerente(conn, gerente_id):
    row = conn.execute(
        "SELECT * FROM gerentes WHERE id = %s AND cliente_id = %s",
        (gerente_id, session["cliente_id"]),
    ).fetchone()
    if not row:
        abort(404)
    gerente = dict(row)
    for campo in ("teto", "peso_a", "peso_b", "peso_c"):
        gerente[campo] = float(gerente[campo])
    return gerente


# ───────────────────────── início / auth ─────────────────────────

@app.route("/")
def index():
    if session.get("is_admin"):
        return redirect(url_for("admin_clientes"))
    if session.get("cliente_id"):
        return redirect(url_for("gerentes"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        tipo, cliente = verificar_login(
            request.form.get("login", "").strip(), request.form.get("senha", "")
        )
        if tipo == "admin":
            session.clear()
            session["is_admin"] = True
            return redirect(url_for("admin_clientes"))
        if tipo == "cliente":
            session.clear()
            session["cliente_id"] = cliente["id"]
            session["cliente_nome"] = cliente["nome"]
            return redirect(url_for("gerentes"))
        flash("Usuário ou senha inválidos.", "erro")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ───────────────────────── admin (super-admin) ─────────────────────────

@app.route("/admin")
@admin_required
def admin_clientes():
    conn = db()
    try:
        clientes = conn.execute(
            """
            SELECT c.*, COUNT(g.id) FILTER (WHERE g.ativo) AS total_gerentes
            FROM clientes c
            LEFT JOIN gerentes g ON g.cliente_id = c.id
            GROUP BY c.id ORDER BY c.nome
            """
        ).fetchall()
    finally:
        conn.close()
    return render_template("admin_clientes.html", clientes=clientes, editar=None)


@app.route("/admin/clientes/novo", methods=["POST"])
@admin_required
def admin_cliente_novo():
    nome = request.form["nome"].strip()
    login_ = request.form["login"].strip()
    senha = request.form["senha"]
    with db() as conn:
        conn.execute(
            "INSERT INTO clientes (nome, login, senha_hash) VALUES (%s, %s, %s)",
            (nome, login_, generate_password_hash(senha, method="pbkdf2:sha256")),
        )
    flash("Cliente criado.")
    return redirect(url_for("admin_clientes"))


@app.route("/admin/clientes/<int:cliente_id>/editar", methods=["GET", "POST"])
@admin_required
def admin_cliente_editar(cliente_id):
    conn = db()
    try:
        if request.method == "POST":
            nome = request.form["nome"].strip()
            login_ = request.form["login"].strip()
            senha = request.form.get("senha", "").strip()
            if senha:
                conn.execute(
                    "UPDATE clientes SET nome=%s, login=%s, senha_hash=%s WHERE id=%s",
                    (nome, login_, generate_password_hash(senha, method="pbkdf2:sha256"), cliente_id),
                )
            else:
                conn.execute(
                    "UPDATE clientes SET nome=%s, login=%s WHERE id=%s",
                    (nome, login_, cliente_id),
                )
            conn.commit()
            flash("Cliente atualizado.")
            return redirect(url_for("admin_clientes"))

        editar = conn.execute("SELECT * FROM clientes WHERE id=%s", (cliente_id,)).fetchone()
        if not editar:
            abort(404)
        clientes = conn.execute(
            """
            SELECT c.*, COUNT(g.id) FILTER (WHERE g.ativo) AS total_gerentes
            FROM clientes c
            LEFT JOIN gerentes g ON g.cliente_id = c.id
            GROUP BY c.id ORDER BY c.nome
            """
        ).fetchall()
    finally:
        conn.close()
    return render_template("admin_clientes.html", clientes=clientes, editar=editar)


@app.route("/admin/clientes/<int:cliente_id>/desativar", methods=["POST"])
@admin_required
def admin_cliente_desativar(cliente_id):
    with db() as conn:
        conn.execute("UPDATE clientes SET ativo = NOT ativo WHERE id=%s", (cliente_id,))
    return redirect(url_for("admin_clientes"))


@app.route("/admin/clientes/<int:cliente_id>/gerentes")
@admin_required
def admin_gerentes(cliente_id):
    conn = db()
    try:
        cliente = conn.execute("SELECT * FROM clientes WHERE id=%s", (cliente_id,)).fetchone()
        if not cliente:
            abort(404)
        gerentes_rows = conn.execute(
            """SELECT g.*, (SELECT COUNT(*) FROM indicadores i WHERE i.gerente_id = g.id) AS total_indicadores
               FROM gerentes g WHERE g.cliente_id=%s ORDER BY g.nome""",
            (cliente_id,),
        ).fetchall()
        todos_clientes = conn.execute(
            "SELECT id, nome FROM clientes WHERE ativo=TRUE ORDER BY nome"
        ).fetchall()
    finally:
        conn.close()
    return render_template(
        "admin_gerentes.html", cliente=cliente, gerentes=gerentes_rows, todos_clientes=todos_clientes
    )


@app.route("/admin/clientes/<int:cliente_id>/gerentes.json")
@admin_required
def admin_gerentes_json(cliente_id):
    conn = db()
    try:
        linhas = conn.execute(
            """SELECT g.id, g.nome, g.loja FROM gerentes g
               WHERE g.cliente_id=%s
                 AND EXISTS (SELECT 1 FROM indicadores i WHERE i.gerente_id = g.id)
               ORDER BY g.nome""",
            (cliente_id,),
        ).fetchall()
    finally:
        conn.close()
    return jsonify([dict(r) for r in linhas])


@app.route("/admin/gerentes/<int:destino_id>/copiar-de/<int:origem_id>", methods=["POST"])
@admin_required
def admin_copiar_indicadores(destino_id, origem_id):
    conn = db()
    try:
        destino = conn.execute("SELECT id FROM gerentes WHERE id=%s", (destino_id,)).fetchone()
        origem = conn.execute("SELECT id FROM gerentes WHERE id=%s", (origem_id,)).fetchone()
        if not destino or not origem:
            abort(404)
        origem_itens = conn.execute(
            "SELECT * FROM indicadores WHERE gerente_id=%s ORDER BY bloco, ordem, id", (origem_id,)
        ).fetchall()
        conn.execute("DELETE FROM indicadores WHERE gerente_id=%s", (destino_id,))
        for linha in origem_itens:
            conn.execute(
                """INSERT INTO indicadores
                       (gerente_id, bloco, ordem, nome, meta, peso, inverso, eh_gatilho,
                        minimo_pct, teto_pct, mult_min, mult_max)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (
                    destino_id, linha["bloco"], linha["ordem"], linha["nome"], linha["meta"], linha["peso"],
                    linha["inverso"], linha["eh_gatilho"], linha["minimo_pct"], linha["teto_pct"],
                    linha["mult_min"], linha["mult_max"],
                ),
            )
        conn.commit()
        return jsonify({"ok": True})
    except Exception as exc:  # noqa: BLE001
        conn.rollback()
        return jsonify({"erro": str(exc)}), 400
    finally:
        conn.close()


# ───────────────────────── gerentes (por cliente) ─────────────────────────

@app.route("/gerentes")
@login_required
def gerentes():
    conn = db()
    try:
        lista = conn.execute(
            "SELECT * FROM gerentes WHERE cliente_id=%s ORDER BY nome",
            (session["cliente_id"],),
        ).fetchall()
    finally:
        conn.close()
    return render_template("gerentes.html", gerentes=lista)


@app.route("/gerentes/novo", methods=["GET", "POST"])
@login_required
def gerente_novo():
    if request.method == "POST":
        with db() as conn:
            conn.execute(
                """INSERT INTO gerentes (cliente_id, nome, loja, teto, peso_a, peso_b, peso_c)
                   VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                (
                    session["cliente_id"],
                    request.form["nome"].strip(),
                    request.form.get("loja", "").strip(),
                    float(request.form["teto"]),
                    float(request.form["peso_a"]),
                    float(request.form["peso_b"]),
                    float(request.form["peso_c"]),
                ),
            )
        flash("Gerente criado.")
        return redirect(url_for("gerentes"))
    return render_template("gerente_form.html", gerente=None)


@app.route("/gerentes/<int:gerente_id>/editar", methods=["GET", "POST"])
@login_required
def gerente_editar(gerente_id):
    conn = db()
    try:
        gerente = carregar_gerente(conn, gerente_id)
        if request.method == "POST":
            conn.execute(
                """UPDATE gerentes SET nome=%s, loja=%s, teto=%s, peso_a=%s, peso_b=%s, peso_c=%s
                   WHERE id=%s AND cliente_id=%s""",
                (
                    request.form["nome"].strip(),
                    request.form.get("loja", "").strip(),
                    float(request.form["teto"]),
                    float(request.form["peso_a"]),
                    float(request.form["peso_b"]),
                    float(request.form["peso_c"]),
                    gerente_id,
                    session["cliente_id"],
                ),
            )
            conn.commit()
            flash("Gerente atualizado.")
            return redirect(url_for("gerentes"))
    finally:
        conn.close()
    return render_template("gerente_form.html", gerente=gerente)


@app.route("/gerentes/<int:gerente_id>/desativar", methods=["POST"])
@login_required
def gerente_desativar(gerente_id):
    with db() as conn:
        carregar_gerente(conn, gerente_id)
        conn.execute(
            "UPDATE gerentes SET ativo = NOT ativo WHERE id=%s AND cliente_id=%s",
            (gerente_id, session["cliente_id"]),
        )
    return redirect(url_for("gerentes"))


# ───────────────────────── indicadores (template vivo) ─────────────────────────

def _linha_indicador(row_or_dict):
    d = dict(row_or_dict)
    return {
        "nome": d["nome"],
        "meta": float(d["meta"]),
        "peso": float(d["peso"]),
        "minimo_pct": float(d["minimo_pct"]),
        "teto_pct": float(d["teto_pct"]),
        "mult_min": float(d["mult_min"]),
        "mult_max": float(d["mult_max"]),
        "inverso": bool(d["inverso"]),
        "eh_gatilho": bool(d["eh_gatilho"]),
    }


@app.route("/gerentes/<int:gerente_id>/indicadores", methods=["GET", "POST"])
@login_required
def indicadores(gerente_id):
    conn = db()
    try:
        gerente = carregar_gerente(conn, gerente_id)

        if request.method == "POST":
            novos = []
            for bloco in ("A", "B", "C"):
                bruto = json.loads(request.form.get(f"itens_{bloco}", "[]"))
                for ordem, item in enumerate(bruto):
                    novos.append(
                        (
                            gerente_id,
                            bloco,
                            ordem,
                            str(item.get("nome", "")).strip() or "Indicador",
                            float(item.get("meta", 0) or 0),
                            float(item.get("peso", 0) or 0),
                            bool(item.get("inverso", False)),
                            bool(item.get("eh_gatilho", False)),
                            float(item.get("minimo_pct", 85) or 0),
                            float(item.get("teto_pct", 110) or 0),
                            float(item.get("mult_min", 0.70) or 0),
                            float(item.get("mult_max", 1.20) or 0),
                        )
                    )
            conn.execute("DELETE FROM indicadores WHERE gerente_id=%s", (gerente_id,))
            for linha in novos:
                conn.execute(
                    """INSERT INTO indicadores
                       (gerente_id, bloco, ordem, nome, meta, peso, inverso, eh_gatilho,
                        minimo_pct, teto_pct, mult_min, mult_max)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    linha,
                )
            conn.commit()
            flash("Indicadores salvos.")
            return redirect(url_for("indicadores", gerente_id=gerente_id))

        linhas = conn.execute(
            "SELECT * FROM indicadores WHERE gerente_id=%s ORDER BY bloco, ordem, id",
            (gerente_id,),
        ).fetchall()
        outros_gerentes = conn.execute(
            """SELECT g.id, g.nome, g.loja FROM gerentes g
               WHERE g.cliente_id=%s AND g.id != %s
                 AND EXISTS (SELECT 1 FROM indicadores i WHERE i.gerente_id = g.id)
               ORDER BY g.nome""",
            (session["cliente_id"], gerente_id),
        ).fetchall()
    finally:
        conn.close()

    por_bloco = {"A": [], "B": [], "C": []}
    for linha in linhas:
        por_bloco[linha["bloco"]].append(_linha_indicador(linha))

    return render_template(
        "indicadores.html",
        gerente=gerente,
        indicadores_json=json.dumps(por_bloco),
        outros_gerentes=outros_gerentes,
    )


@app.route("/gerentes/<int:gerente_id>/indicadores.json")
@login_required
def indicadores_json(gerente_id):
    conn = db()
    try:
        carregar_gerente(conn, gerente_id)  # garante que pertence ao cliente logado
        linhas = conn.execute(
            "SELECT * FROM indicadores WHERE gerente_id=%s ORDER BY bloco, ordem, id",
            (gerente_id,),
        ).fetchall()
    finally:
        conn.close()

    por_bloco = {"A": [], "B": [], "C": []}
    for linha in linhas:
        por_bloco[linha["bloco"]].append(_linha_indicador(linha))
    return jsonify(por_bloco)


# ───────────────────────── calculadora do mês ─────────────────────────

def carregar_calculo_mes(conn, gerente, gerente_id, mes):
    """Carrega os dados do mês (salvo) ou monta um rascunho a partir do
    template vivo de indicadores (ainda não salvo). Usado tanto pela tela
    da calculadora quanto pela exportação em Excel."""
    calculo = conn.execute(
        "SELECT * FROM calculos WHERE gerente_id=%s AND mes=%s", (gerente_id, mes)
    ).fetchone()

    if calculo:
        gerente_calc = dict(gerente)
        gerente_calc.update(
            {
                "teto": float(calculo["teto"]),
                "peso_a": float(calculo["peso_a"]),
                "peso_b": float(calculo["peso_b"]),
                "peso_c": float(calculo["peso_c"]),
            }
        )
        itens_rows = conn.execute(
            "SELECT * FROM calculo_itens WHERE calculo_id=%s ORDER BY bloco, ordem, id",
            (calculo["id"],),
        ).fetchall()
        itens_por_bloco = {"A": [], "B": [], "C": []}
        for row in itens_rows:
            item = _linha_indicador(row)
            item["realizado"] = float(row["realizado"])
            itens_por_bloco[row["bloco"]].append(item)
        ajustes_rows = conn.execute(
            "SELECT nome, valor FROM calculo_ajustes WHERE calculo_id=%s ORDER BY ordem, id",
            (calculo["id"],),
        ).fetchall()
        ajustes = [{"nome": r["nome"], "valor": float(r["valor"])} for r in ajustes_rows]
        ja_salvo = True
        atualizado_em = calculo["atualizado_em"].strftime("%d/%m/%Y %H:%M")
    else:
        gerente_calc = dict(gerente)
        linhas = conn.execute(
            "SELECT * FROM indicadores WHERE gerente_id=%s ORDER BY bloco, ordem, id",
            (gerente_id,),
        ).fetchall()
        itens_por_bloco = {"A": [], "B": [], "C": []}
        for linha in linhas:
            item = _linha_indicador(linha)
            item["realizado"] = 0.0
            itens_por_bloco[linha["bloco"]].append(item)
        ajustes = []
        ja_salvo = False
        atualizado_em = None

    return {
        "gerente_calc": gerente_calc,
        "itens_por_bloco": itens_por_bloco,
        "ajustes": ajustes,
        "ja_salvo": ja_salvo,
        "atualizado_em": atualizado_em,
    }


@app.route("/gerentes/<int:gerente_id>/mes")
@login_required
def calculadora_mes_atual(gerente_id):
    return redirect(url_for("calculadora", gerente_id=gerente_id, mes=date.today().strftime("%Y-%m")))


@app.route("/gerentes/<int:gerente_id>/mes/<mes>")
@login_required
def calculadora(gerente_id, mes):
    conn = db()
    try:
        gerente = carregar_gerente(conn, gerente_id)
        dados = carregar_calculo_mes(conn, gerente, gerente_id, mes)
    finally:
        conn.close()

    gerente_calc = dados["gerente_calc"]
    return render_template(
        "calculadora.html",
        gerente=gerente_calc,
        mes=mes,
        mes_label=mes_label(mes),
        ja_salvo=dados["ja_salvo"],
        atualizado_em=dados["atualizado_em"],
        gerente_json=json.dumps(
            {
                "teto": gerente_calc["teto"],
                "peso_a": gerente_calc["peso_a"],
                "peso_b": gerente_calc["peso_b"],
                "peso_c": gerente_calc["peso_c"],
            }
        ),
        itens_json=json.dumps(dados["itens_por_bloco"]),
        ajustes_json=json.dumps(dados["ajustes"]),
    )


NOME_BLOCO = {"A": "Bloco A — Financeiro", "B": "Bloco B — Processos", "C": "Bloco C — Pessoas"}


@app.route("/gerentes/<int:gerente_id>/mes/<mes>/exportar.xlsx")
@login_required
def calculadora_exportar(gerente_id, mes):
    conn = db()
    try:
        gerente = carregar_gerente(conn, gerente_id)
        dados = carregar_calculo_mes(conn, gerente, gerente_id, mes)
    finally:
        conn.close()

    gerente_calc = dados["gerente_calc"]
    resultado = calcular_mes(gerente_calc, dados["itens_por_bloco"], dados["ajustes"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Premiação"

    negrito = Font(bold=True)
    negrito_grande = Font(bold=True, size=13)
    fmt_moeda = "R$ #,##0.00"
    fmt_pct = "0%"

    ws.append([f"{gerente_calc['nome']} — {gerente_calc['loja']}"])
    ws["A1"].font = negrito_grande
    ws.append([mes_label(mes)])
    ws.append([])

    linha = 4
    for bloco in ("A", "B", "C"):
        ws.append([NOME_BLOCO[bloco]])
        ws.cell(row=linha, column=1).font = negrito
        linha += 1
        ws.append(["Indicador", "Meta", "Peso %", "Realizado", "% Atingido", "Prêmio calculado"])
        for cel in ws[linha]:
            cel.font = negrito
        linha += 1

        itens = dados["itens_por_bloco"][bloco]
        premios = resultado["premios"][bloco]
        for item, premio in zip(itens, premios):
            ating = atingimento(item["realizado"], item["meta"], item["inverso"])
            ws.append([item["nome"], item["meta"], item["peso"] / 100, item["realizado"], ating, premio])
            ws.cell(row=linha, column=3).number_format = fmt_pct
            ws.cell(row=linha, column=5).number_format = fmt_pct
            ws.cell(row=linha, column=6).number_format = fmt_moeda
            linha += 1

        ws.append([f"Total {NOME_BLOCO[bloco]}", "", "", "", "", resultado[f"total_{bloco.lower()}"]])
        for cel in ws[linha]:
            cel.font = negrito
        ws.cell(row=linha, column=6).number_format = fmt_moeda
        linha += 1
        ws.append([])
        linha += 1

    ws.append(["Ajustes (prêmios extras / penalidades)"])
    ws.cell(row=linha, column=1).font = negrito
    linha += 1
    ws.append(["Descrição", "Valor"])
    for cel in ws[linha]:
        cel.font = negrito
    linha += 1
    for ajuste in dados["ajustes"]:
        ws.append([ajuste["nome"], ajuste["valor"]])
        ws.cell(row=linha, column=2).number_format = fmt_moeda
        linha += 1
    ws.append(["Soma dos ajustes", resultado["soma_ajustes"]])
    for cel in ws[linha]:
        cel.font = negrito
    ws.cell(row=linha, column=2).number_format = fmt_moeda
    linha += 2

    ws.append(["Total a pagar", resultado["total"]])
    for cel in ws[linha]:
        cel.font = negrito_grande
    ws.cell(row=linha, column=2).number_format = fmt_moeda

    ws.column_dimensions["A"].width = 32
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f"premiacao_{gerente_calc['nome']}_{mes}.xlsx".replace(" ", "_")
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/gerentes/<int:gerente_id>/mes/<mes>/salvar", methods=["POST"])
@login_required
def calculadora_salvar(gerente_id, mes):
    payload = request.get_json(silent=True) or {}
    itens_enviados = payload.get("itens", {})
    ajustes_enviados = payload.get("ajustes", [])

    conn = db()
    try:
        gerente = carregar_gerente(conn, gerente_id)

        itens_por_bloco = {"A": [], "B": [], "C": []}
        for bloco in ("A", "B", "C"):
            for ordem, item in enumerate(itens_enviados.get(bloco, [])):
                itens_por_bloco[bloco].append(
                    {
                        "ordem": ordem,
                        "nome": str(item.get("nome", "")).strip() or "Indicador",
                        "meta": float(item.get("meta", 0) or 0),
                        "peso": float(item.get("peso", 0) or 0),
                        "inverso": bool(item.get("inverso", False)),
                        "eh_gatilho": bool(item.get("eh_gatilho", False)),
                        "minimo_pct": float(item.get("minimo_pct", 85) or 0),
                        "teto_pct": float(item.get("teto_pct", 110) or 0),
                        "mult_min": float(item.get("mult_min", 0.70) or 0),
                        "mult_max": float(item.get("mult_max", 1.20) or 0),
                        "realizado": float(item.get("realizado", 0) or 0),
                    }
                )

        ajustes = []
        for ordem, ajuste in enumerate(ajustes_enviados):
            nome = str(ajuste.get("nome", "")).strip()
            valor = float(ajuste.get("valor", 0) or 0)
            if not nome and valor == 0:
                continue
            ajustes.append({"ordem": ordem, "nome": nome or "Ajuste", "valor": valor})

        resultado = calcular_mes(gerente, itens_por_bloco, ajustes)

        calculo = conn.execute(
            "SELECT id FROM calculos WHERE gerente_id=%s AND mes=%s", (gerente_id, mes)
        ).fetchone()

        if calculo:
            calculo_id = calculo["id"]
            conn.execute(
                """UPDATE calculos SET teto=%s, peso_a=%s, peso_b=%s, peso_c=%s,
                       total_a=%s, total_b=%s, total_c=%s, total=%s, atualizado_em=now()
                   WHERE id=%s""",
                (
                    gerente["teto"], gerente["peso_a"], gerente["peso_b"], gerente["peso_c"],
                    resultado["total_a"], resultado["total_b"], resultado["total_c"],
                    resultado["total"], calculo_id,
                ),
            )
            conn.execute("DELETE FROM calculo_itens WHERE calculo_id=%s", (calculo_id,))
            conn.execute("DELETE FROM calculo_ajustes WHERE calculo_id=%s", (calculo_id,))
        else:
            row = conn.execute(
                """INSERT INTO calculos
                       (gerente_id, mes, teto, peso_a, peso_b, peso_c,
                        total_a, total_b, total_c, total)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (
                    gerente_id, mes, gerente["teto"], gerente["peso_a"], gerente["peso_b"], gerente["peso_c"],
                    resultado["total_a"], resultado["total_b"], resultado["total_c"], resultado["total"],
                ),
            ).fetchone()
            calculo_id = row["id"]

        for bloco in ("A", "B", "C"):
            premios = resultado["premios"][bloco]
            for item, premio in zip(itens_por_bloco[bloco], premios):
                conn.execute(
                    """INSERT INTO calculo_itens
                           (calculo_id, bloco, ordem, nome, meta, peso, inverso, eh_gatilho,
                            minimo_pct, teto_pct, mult_min, mult_max, realizado, premio_calculado)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (
                        calculo_id, bloco, item["ordem"], item["nome"], item["meta"], item["peso"],
                        item["inverso"], item["eh_gatilho"], item["minimo_pct"], item["teto_pct"],
                        item["mult_min"], item["mult_max"], item["realizado"], premio,
                    ),
                )

        for ajuste in ajustes:
            conn.execute(
                "INSERT INTO calculo_ajustes (calculo_id, ordem, nome, valor) VALUES (%s,%s,%s,%s)",
                (calculo_id, ajuste["ordem"], ajuste["nome"], ajuste["valor"]),
            )

        conn.commit()
        return jsonify({"ok": True, "total": resultado["total"]})
    except Exception as exc:  # noqa: BLE001
        conn.rollback()
        return jsonify({"erro": str(exc)}), 400
    finally:
        conn.close()


# ───────────────────────── histórico ─────────────────────────

@app.route("/gerentes/<int:gerente_id>/historico")
@login_required
def historico(gerente_id):
    conn = db()
    try:
        gerente = carregar_gerente(conn, gerente_id)
        calculos = conn.execute(
            "SELECT * FROM calculos WHERE gerente_id=%s ORDER BY mes DESC", (gerente_id,)
        ).fetchall()
    finally:
        conn.close()
    return render_template("historico.html", gerente=gerente, calculos=calculos)


if __name__ == "__main__":
    app.run(debug=True)
